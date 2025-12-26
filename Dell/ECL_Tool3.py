import re
import csv
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from typing import List, Dict, Tuple, Optional, Any, Set

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.hyperlink import Hyperlink
    import openpyxl.styles
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# --- Helper Functions ---

def get_refdes(pin_name: str) -> Optional[str]:
    if not pin_name: return None
    pin_name = pin_name.replace('*', '')
    if pin_name.upper() == 'T': return None
    if pin_name.upper().startswith('VIA'): return None
    if '.' in pin_name:
        return pin_name.split('.')[0]
    return pin_name

def is_component_pin(name: str) -> bool:
    n = name.replace('*', '').upper()
    if n.startswith('VIA'): return False
    if n == 'T': return False
    return True

# --- Regex Patterns ---
re_item = re.compile(r'^\s*([A-Za-z0-9_\.\/\*]+)\s+([\-0-9\.]+)\s+([\-0-9\.]+)\s+([LDV])?\s*([0-9\.]+)\s*(.*)?$')
re_total = re.compile(r'^\s*TOTAL\s+(\d+)\s+VIA\(S\)\s+([0-9\.]+)\s+mils')

def parse_file(path: str) -> List[Dict]:
    nets = []
    current_net_name = None
    items = [] 

    with open(path, 'r', encoding='utf-8', errors='ignore') as f:
        lines = [ln.strip() for ln in f]

    for ln in lines:
        if not ln: continue
        if ln.startswith('|') or ln.startswith('ECL ') or ln.startswith('Page') or ln.startswith('C:/'): continue
        if ln.startswith('dimensions') or ln.startswith('refdes'): continue
        if ln.startswith('net name') or ln.startswith('End of ECL') or ln.startswith('total path length'): continue
        if ' - - - ' in ln: continue
        if ln.startswith('\x0c'): continue

        m_total = re_total.match(ln)
        if m_total:
            if current_net_name and items:
                final_total_len = float(m_total.group(2))
                start_pin = items[0][0].replace('*', '')

                full_path_net = {
                    'source_file': os.path.basename(path),
                    'net_name': current_net_name,
                    'start_pin': start_pin,
                    'items': items,
                    'total': final_total_len,
                    'via_count_reported': int(m_total.group(1))
                }
                nets.append(full_path_net)
                
            current_net_name = None
            items = []
            continue

        if not re.search(r'[\-0-9]+\.[0-9]+', ln):
            current_net_name = ln.strip()
            items = []
            continue

        m_item = re_item.match(ln)
        if m_item and current_net_name:
            name = m_item.group(1)
            type_char = m_item.group(4)
            length = float(m_item.group(5))
            layer = m_item.group(6) if m_item.group(6) else ''
            items.append((name, length, layer, type_char))
            continue

    return nets

def convert_net_to_segments(n: Dict) -> Dict:
    items = n['items']
    if not items: return {}
    
    start_pin = n['start_pin']
    end_pin_name, end_total_len, end_layer, _ = items[-1]
    end_pin = end_pin_name.replace('*', '')
    
    raw_segments = []
    via_count_calc = 0
    prev_len = 0.0
    
    for i in range(1, len(items)):
        curr_name, curr_len, curr_layer, _ = items[i]
        curr_name_clean = curr_name.replace('*', '')
        
        seg_len = curr_len - prev_len
        seg_layer = curr_layer.strip()
        
        next_conn = None
        if i < len(items) - 1:
            if curr_name_clean.upper().startswith('VIA'):
                next_conn = 'VIA'
                via_count_calc += 1
            elif curr_name_clean.upper() == 'T':
                next_conn = None 
            else:
                next_conn = get_refdes(curr_name_clean) 
        
        if i == len(items) - 1:
            if curr_name_clean.upper().startswith('VIA'):
                via_count_calc += 1
        
        raw_segments.append({
            'layer': seg_layer,
            'length': seg_len,
            'next_conn': next_conn
        })
        prev_len = curr_len

    final_segments = []
    for i, seg in enumerate(raw_segments):
        if seg['length'] < 0.01:
            if final_segments and seg['next_conn']:
                final_segments[-1]['next_conn'] = seg['next_conn']
        else:
            final_segments.append(seg)

    return {
        'source_file': n['source_file'],
        'net_name': n['net_name'],
        'start_pin': start_pin,
        'end_pin': end_pin,
        'segments': final_segments,
        'via_count': via_count_calc,
        'total': end_total_len
    }

def merge_passive_nets(rows: List[Dict]) -> List[Dict]:
    """
    標準一對一合併邏輯 (Stable Version)。
    """
    def get_clean_name(name):
        if '_C_' in name: return name.replace('_C_', '_')
        if '_R_' in name: return name.replace('_R_', '_')
        return name

    while True:
        merged_any = False
        comp_map = {}
        
        for i, row in enumerate(rows):
            s_pin = row['start_pin']
            e_pin = row['end_pin']
            for pin, ptype in [(s_pin, 'Start'), (e_pin, 'End')]:
                ref = get_refdes(pin)
                if ref and (ref.startswith('C') or ref.startswith('R') or ref.startswith('L')):
                    if ref not in comp_map:
                        comp_map[ref] = []
                    comp_map[ref].append((i, pin, ptype))
        
        indices_to_remove = set()
        new_rows = []
        
        for ref in sorted(comp_map.keys()):
            entries = comp_map[ref]
            valid_entries = [e for e in entries if e[0] not in indices_to_remove]
            
            if len(valid_entries) != 2: continue
            (idx1, pin1, type1), (idx2, pin2, type2) = valid_entries
            if pin1 == pin2: continue
            
            row1 = rows[idx1]
            row2 = rows[idx2]
            
            indices_to_remove.add(idx1)
            indices_to_remove.add(idx2)
            merged_any = True
            
            # --- Merge Logic ---
            segs1 = row1['segments'][:]
            start_pin_final = row1['start_pin']
            if type1 == 'Start':
                rev_segs = []
                conns = [s['next_conn'] for s in segs1[:-1]]
                bodies = [{'layer': s['layer'], 'length': s['length']} for s in segs1]
                bodies.reverse()
                conns.reverse()
                for x in range(len(bodies)):
                    s = bodies[x]
                    nxt = conns[x] if x < len(conns) else None
                    s['next_conn'] = nxt
                    rev_segs.append(s)
                segs1 = rev_segs
                start_pin_final = row1['end_pin']
            
            segs2 = row2['segments'][:]
            end_pin_final = row2['end_pin']
            if type2 == 'End':
                rev_segs = []
                conns = [s['next_conn'] for s in segs2[:-1]]
                bodies = [{'layer': s['layer'], 'length': s['length']} for s in segs2]
                bodies.reverse()
                conns.reverse()
                for x in range(len(bodies)):
                    s = bodies[x]
                    nxt = conns[x] if x < len(conns) else None
                    s['next_conn'] = nxt
                    rev_segs.append(s)
                segs2 = rev_segs
                end_pin_final = row2['start_pin']
            
            if segs1:
                segs1[-1]['next_conn'] = ref
            
            merged_segs = segs1 + segs2
            total_len = sum(s['length'] for s in merged_segs)
            
            c1 = row1['via_count']
            c2 = row2['via_count']
            new_via_count = c1 + c2
            
            n1 = row1['net_name']
            n2 = row2['net_name']
            clean1 = get_clean_name(n1)
            clean2 = get_clean_name(n2)
            final_name = clean2 if len(clean2) < len(n2) else clean1

            new_row = {
                'source_file': row1['source_file'],
                'net_name': final_name,
                'start_pin': start_pin_final,
                'end_pin': end_pin_final,
                'segments': merged_segs,
                'via_count': new_via_count,
                'total': total_len
            }
            new_rows.append(new_row)
        
        if not merged_any:
            break
            
        final_list = [r for i, r in enumerate(rows) if i not in indices_to_remove]
        final_list.extend(new_rows)
        rows = final_list
    
    return rows

def collect_and_sort_layers(grouped_data: Dict) -> List[str]:
    """
    收集所有 Layer 並按照 TOP -> L1 -> L2... -> BOTTOM 順序排序。
    """
    all_layers = set()
    for (ref1, ref2), rows in grouped_data.items():
        for row in rows:
            for seg in row['segments']:
                layer = seg['layer']
                if layer:
                    all_layers.add(layer.strip())
    
    def layer_sort_key(name):
        u = name.upper()
        if 'TOP' in u: return -1000
        if 'BOTTOM' in u or 'BOT' in u: return 1000
        nums = re.findall(r'\d+', name)
        if nums: return int(nums[0])
        return 0
        
    sorted_layers = sorted(list(all_layers), key=layer_sort_key)
    return sorted_layers

def combine_to_excel(input_files: List[str], output_file: str, include_source: bool = True):
    if not HAS_OPENPYXL:
        raise ImportError("The 'openpyxl' library is required for Excel export. Please install it via 'pip install openpyxl'.")

    all_raw_nets = []
    for fp in input_files:
        if not os.path.exists(fp):
            print(f'[WARN] file not found: {fp}')
            continue
        nets = parse_file(fp)
        all_raw_nets.extend(nets)

    if not all_raw_nets:
        raise RuntimeError('No net resolved.')

    segment_rows = [convert_net_to_segments(n) for n in all_raw_nets]
    merged_rows = merge_passive_nets(segment_rows)
    
    # --- Grouping ---
    grouped_data = {} 
    for r in merged_rows:
        s_ref = get_refdes(r['start_pin']) or "Unknown"
        e_ref = get_refdes(r['end_pin']) or "Unknown"
        refs = sorted([s_ref, e_ref])
        key = (refs[0], refs[1])
        if key not in grouped_data:
            grouped_data[key] = []
        grouped_data[key].append(r)
    
    # --- Collect & Sort Layers ---
    sorted_layers = collect_and_sort_layers(grouped_data)
    
    # --- Write Excel ---
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # 1. Summary Sheet
    summary_ws = wb.create_sheet(title='Summary', index=0)
    summary_ws['A1'] = 'ECL Report Summary'
    summary_ws['A1'].font = openpyxl.styles.Font(size=14, bold=True)
    summary_ws['A3'] = 'Sheet Name'
    summary_ws['B3'] = 'Component Pair'
    summary_ws['C3'] = 'Net Count'
    for cell in ['A3', 'B3', 'C3']:
        summary_ws[cell].font = openpyxl.styles.Font(bold=True)
    summary_row_idx = 4
    
    # 2. Stack up Sheet
    stackup_ws = wb.create_sheet(title='Stack up', index=1)
    
    stackup_ws['A1'] = 'Layer'
    stackup_ws['B1'] = '4G'
    stackup_ws['C1'] = '8G'
    stackup_ws['D1'] = '16G'
    
    for col in ['A', 'B', 'C', 'D']:
        cell = stackup_ws[f'{col}1']
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))

    stackup_row = 2
    for layer in sorted_layers:
        stackup_ws[f'A{stackup_row}'] = layer
        stackup_row += 1
    
    extra_items = ['VIA', 'Cap', 'CONN']
    for item in extra_items:
        cell = stackup_ws[f'A{stackup_row}']
        cell.value = item
        cell.font = openpyxl.styles.Font(bold=True, italic=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        stackup_row += 1
    
    stackup_ws.column_dimensions['A'].width = 25
    stackup_ws.column_dimensions['B'].width = 15
    stackup_ws.column_dimensions['C'].width = 15
    stackup_ws.column_dimensions['D'].width = 15

    # --- Helper function for net name check ---
    def should_have_loss_calc(net_name):
        name_upper = net_name.strip().upper()
        return name_upper.startswith('PE') or name_upper.startswith('XGMI') or name_upper.startswith('UPI')
        
    # 3. Data Sheets
    for (ref1, ref2), rows in sorted(grouped_data.items()):
        sheet_name = f"{ref1}_{ref2}"
        invalid_chars = '[]:*?/\\' 
        for ch in invalid_chars:
            sheet_name = sheet_name.replace(ch, '_')
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        
        ws = wb.create_sheet(title=sheet_name)
        
        has_loss_calc_net = any(should_have_loss_calc(r['net_name']) for r in rows)
        
        max_segments = 0
        for r in rows:
            max_segments = max(max_segments, len(r['segments']))
            
        header = ['net_name', 'Start Pin']
        if include_source: header.insert(0, 'source_file')
        
        for i in range(1, max_segments + 1):
            header += [f'Layer_{i}', f'Length_{i}']
            if i < max_segments:
                header += [f'VIA_{i}']
        header += ['End Pin', 'total length', 'via count']
        
        if has_loss_calc_net:
            header += ['Loss at 4G', 'Loss at 8G', 'Loss at 16G']
        
        ws.append(header)
        
        rows.sort(key=lambda x: x['net_name'])
        
        current_row_idx = 2
        for r in rows:
            row_data = []
            if include_source: row_data.append(r['source_file'])
            row_data += [r['net_name'], r['start_pin']]
            
            segs = r['segments']
            
            formula_components_4g = []
            formula_components_8g = []
            formula_components_16g = []
            
            is_loss_net = should_have_loss_calc(r['net_name'])
            
            base_col_idx = 4 if include_source else 3
            
            for i in range(max_segments):
                if i < len(segs):
                    s = segs[i]
                    row_data += [s['layer'], round(s['length'], 2)]
                    
                    if is_loss_net:
                        layer_col_letter = get_column_letter(base_col_idx + i*3)
                        len_col_letter = get_column_letter(base_col_idx + i*3 + 1)
                        
                        term_4g = f"(({len_col_letter}{current_row_idx}/1000)*VLOOKUP({layer_col_letter}{current_row_idx},'Stack up'!$A:$D,2,0))"
                        term_8g = f"(({len_col_letter}{current_row_idx}/1000)*VLOOKUP({layer_col_letter}{current_row_idx},'Stack up'!$A:$D,3,0))"
                        term_16g = f"(({len_col_letter}{current_row_idx}/1000)*VLOOKUP({layer_col_letter}{current_row_idx},'Stack up'!$A:$D,4,0))"
                        
                        formula_components_4g.append(term_4g)
                        formula_components_8g.append(term_8g)
                        formula_components_16g.append(term_16g)
                    
                    if i < max_segments - 1:
                        conn = s['next_conn']
                        row_data += [conn if conn else '']
                else:
                    row_data += ['', '']
                    if i < max_segments - 1:
                        row_data += ['']
            
            row_data += [r['end_pin'], round(r['total'], 2), r['via_count']]
            
            if has_loss_calc_net:
                if is_loss_net:
                    via_cnt = r['via_count']
                    
                    cap_cnt = 0
                    for s in segs:
                        if s['next_conn'] and s['next_conn'].startswith('C'):
                            cap_cnt += 1
                            
                    conn_cnt = 0
                    s_ref = get_refdes(r['start_pin'])
                    e_ref = get_refdes(r['end_pin'])
                    if s_ref and (s_ref.startswith('J') or s_ref.startswith('CONN')): conn_cnt += 1
                    if e_ref and (e_ref.startswith('J') or e_ref.startswith('CONN')): conn_cnt += 1
                    
                    if via_cnt > 0:
                        f_via_4g = f"({via_cnt}*VLOOKUP(\"VIA\",'Stack up'!$A:$D,2,0))"
                        f_via_8g = f"({via_cnt}*VLOOKUP(\"VIA\",'Stack up'!$A:$D,3,0))"
                        f_via_16g = f"({via_cnt}*VLOOKUP(\"VIA\",'Stack up'!$A:$D,4,0))"
                        formula_components_4g.append(f_via_4g)
                        formula_components_8g.append(f_via_8g)
                        formula_components_16g.append(f_via_16g)
                        
                    if cap_cnt > 0:
                        f_cap_4g = f"({cap_cnt}*VLOOKUP(\"Cap\",'Stack up'!$A:$D,2,0))"
                        f_cap_8g = f"({cap_cnt}*VLOOKUP(\"Cap\",'Stack up'!$A:$D,3,0))"
                        f_cap_16g = f"({cap_cnt}*VLOOKUP(\"Cap\",'Stack up'!$A:$D,4,0))"
                        formula_components_4g.append(f_cap_4g)
                        formula_components_8g.append(f_cap_8g)
                        formula_components_16g.append(f_cap_16g)
                        
                    if conn_cnt > 0:
                        f_conn_4g = f"({conn_cnt}*VLOOKUP(\"CONN\",'Stack up'!$A:$D,2,0))"
                        f_conn_8g = f"({conn_cnt}*VLOOKUP(\"CONN\",'Stack up'!$A:$D,3,0))"
                        f_conn_16g = f"({conn_cnt}*VLOOKUP(\"CONN\",'Stack up'!$A:$D,4,0))"
                        formula_components_4g.append(f_conn_4g)
                        formula_components_8g.append(f_conn_8g)
                        formula_components_16g.append(f_conn_16g)
                    
                    final_f_4g = "=" + "+".join(formula_components_4g) if formula_components_4g else ""
                    final_f_8g = "=" + "+".join(formula_components_8g) if formula_components_8g else ""
                    final_f_16g = "=" + "+".join(formula_components_16g) if formula_components_16g else ""
                    
                    row_data += [final_f_4g, final_f_8g, final_f_16g]
                else:
                    row_data += ['', '', '']
                
            ws.append(row_data)
            current_row_idx += 1
        
        cell = summary_ws[f'A{summary_row_idx}']
        cell.value = sheet_name
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.font = openpyxl.styles.Font(color="0563C1", underline="single")
        summary_ws[f'B{summary_row_idx}'] = f"{ref1} <-> {ref2}"
        summary_ws[f'C{summary_row_idx}'] = len(rows)
        summary_row_idx += 1
    
    summary_ws.column_dimensions['A'].width = 30
    summary_ws.column_dimensions['B'].width = 35

    wb.save(output_file)

def run_gui():
    root = tk.Tk()
    root.withdraw()
    
    if not HAS_OPENPYXL:
        messagebox.showerror('Missing Library', "Please install 'openpyxl' to use Excel export feature.\nRun: pip install openpyxl")
        return

    messagebox.showinfo('Select File', 'Please select one or more ECL report files (txt)')
    file_paths = filedialog.askopenfilenames(
        title='Choose the ECL report file',
        filetypes=[('Text files', '*.txt'), ('All files', '*.*')]
    )
    if not file_paths: return
    
    ofile = filedialog.asksaveasfilename(
        title='output Excel file',
        defaultextension='.xlsx',
        filetypes=[('Excel files', '*.xlsx')],
        initialfile='ecl_reports_combined.xlsx'
    )
    if not ofile: return

    try:
        combine_to_excel(list(file_paths), ofile, include_source=True)
        messagebox.showinfo('Completed', f'Exported: {ofile}')
    except Exception as e:
        messagebox.showerror('Error', f'Processing failure: {e}')
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    try:
        run_gui()
    except tk.TclError:
        print('GUI not available.')
