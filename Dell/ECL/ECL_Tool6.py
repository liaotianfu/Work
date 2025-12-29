import re
import csv
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from typing import List, Dict, Tuple, Optional, Any, Set

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, quote_sheetname
    from openpyxl.worksheet.hyperlink import Hyperlink
    from openpyxl.worksheet.datavalidation import DataValidation
    import openpyxl.styles
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# --- Helper Functions ---

def get_refdes(pin_name: str) -> Optional[str]:
    if not pin_name: return None
    pin_name = pin_name.replace('*', '')
    if pin_name.upper() == 'T': return None
    if pin_name.upper().startswith('VIA'): return None
    if '.' in pin_name:
        return pin_name.split('.')[0]
    return pin_name


def is_component_pin(name: str) -> bool:
    n = name.replace('*', '').upper()
    if n.startswith('VIA'): return False
    if n == 'T': return False
    return True


# --- Regex Patterns ---
re_item = re.compile(r'^\s*([A-Za-z0-9_\.\/\*]+)\s+([\-0-9\.]+)\s+([\-0-9\.]+)\s+([LBDV])?\s*([0-9\.]+)\s*(.*)?$')
re_total = re.compile(r'^\s*TOTAL\s+(\d+)\s+VIA\(S\)\s+([0-9\.]+)\s+mils')


def parse_file(path: str) -> List[Dict]:
    nets = []
    current_net_name = None
    items = [] 

    with open(path, 'r', encoding='utf-8', errors='ignore') as f:
        lines = [ln.strip() for ln in f]

    for ln in lines:
        if not ln: continue
        if ln.startswith('|') or ln.startswith('ECL ') or ln.startswith('Page') or ln.startswith('C:/'): continue
        if ln.startswith('dimensions') or ln.startswith('refdes'): continue
        if ln.startswith('net name') or ln.startswith('End of ECL') or ln.startswith('total path length'): continue
        if ' - - - ' in ln: continue
        if ln.startswith('\x0c'): continue

        m_total = re_total.match(ln)
        if m_total:
            if current_net_name and items:
                final_total_len = float(m_total.group(2))
                start_pin = items[0][0].replace('*', '')

                full_path_net = {
                    'source_file': os.path.basename(path),
                    'net_name': current_net_name,
                    'start_pin': start_pin,
                    'items': items,
                    'total': final_total_len,
                    'via_count_reported': int(m_total.group(1))
                }
                nets.append(full_path_net)
                
            current_net_name = None
            items = []
            continue

        if not re.search(r'[\-0-9]+\.[0-9]+', ln):
            current_net_name = ln.strip()
            items = []
            continue

        m_item = re_item.match(ln)
        if m_item and current_net_name:
            name = m_item.group(1)
            type_char = m_item.group(4)
            length = float(m_item.group(5))
            layer = m_item.group(6) if m_item.group(6) else ''
            items.append((name, length, layer, type_char))
            continue

    return nets


def convert_net_to_segments(n: Dict) -> Dict:
    items = n['items']
    if not items: return {}
    
    start_pin = n['start_pin']
    end_pin_name, end_total_len, end_layer, _ = items[-1]
    end_pin = end_pin_name.replace('*', '')
    
    raw_segments = []
    via_count_calc = 0
    prev_len = 0.0
    
    for i in range(1, len(items)):
        curr_name, curr_len, curr_layer, _ = items[i]
        curr_name_clean = curr_name.replace('*', '')
        
        seg_len = curr_len - prev_len
        seg_layer = curr_layer.strip()
        
        next_conn = None
        if i < len(items) - 1:
            if curr_name_clean.upper().startswith('VIA'):
                next_conn = 'VIA'
                via_count_calc += 1
            elif curr_name_clean.upper() == 'T':
                next_conn = None 
            else:
                next_conn = get_refdes(curr_name_clean) 
        
        if i == len(items) - 1:
            if curr_name_clean.upper().startswith('VIA'):
                via_count_calc += 1
        
        raw_segments.append({
            'layer': seg_layer,
            'length': seg_len,
            'next_conn': next_conn
        })
        prev_len = curr_len

    final_segments = []
    for i, seg in enumerate(raw_segments):
        if seg['length'] < 0.02:
            if final_segments and seg['next_conn']:
                final_segments[-1]['next_conn'] = seg['next_conn']
        else:
            final_segments.append(seg)

    return {
        'source_file': n['source_file'],
        'net_name': n['net_name'],
        'start_pin': start_pin,
        'end_pin': end_pin,
        'segments': final_segments,
        'via_count': via_count_calc,
        'total': end_total_len
    }


def merge_passive_nets(rows: List[Dict]) -> List[Dict]:
    """
    標準一對一合併邏輯 (Stable Version)。
    """
    def get_clean_name(name):
        if '_C_' in name: return name.replace('_C_', '_')
        if '_R_' in name: return name.replace('_R_', '_')
        return name

    while True:
        merged_any = False
        comp_map = {}
        
        for i, row in enumerate(rows):
            s_pin = row['start_pin']
            e_pin = row['end_pin']
            for pin, ptype in [(s_pin, 'Start'), (e_pin, 'End')]:
                ref = get_refdes(pin)
                if ref and (ref.startswith('C') or ref.startswith('R') or ref.startswith('L')):
                    if ref not in comp_map:
                        comp_map[ref] = []
                    comp_map[ref].append((i, pin, ptype))
        
        indices_to_remove = set()
        new_rows = []
        
        for ref in sorted(comp_map.keys()):
            entries = comp_map[ref]
            valid_entries = [e for e in entries if e[0] not in indices_to_remove]
            
            if len(valid_entries) != 2: continue
            (idx1, pin1, type1), (idx2, pin2, type2) = valid_entries
            if pin1 == pin2: continue
            
            row1 = rows[idx1]
            row2 = rows[idx2]
            
            indices_to_remove.add(idx1)
            indices_to_remove.add(idx2)
            merged_any = True
            
            # --- Merge Logic ---
            segs1 = row1['segments'][:]
            start_pin_final = row1['start_pin']
            if type1 == 'Start':
                rev_segs = []
                conns = [s['next_conn'] for s in segs1[:-1]]
                bodies = [{'layer': s['layer'], 'length': s['length']} for s in segs1]
                bodies.reverse()
                conns.reverse()
                for x in range(len(bodies)):
                    s = bodies[x]
                    nxt = conns[x] if x < len(conns) else None
                    s['next_conn'] = nxt
                    rev_segs.append(s)
                segs1 = rev_segs
                start_pin_final = row1['end_pin']
            
            segs2 = row2['segments'][:]
            end_pin_final = row2['end_pin']
            if type2 == 'End':
                rev_segs = []
                conns = [s['next_conn'] for s in segs2[:-1]]
                bodies = [{'layer': s['layer'], 'length': s['length']} for s in segs2]
                bodies.reverse()
                conns.reverse()
                for x in range(len(bodies)):
                    s = bodies[x]
                    nxt = conns[x] if x < len(conns) else None
                    s['next_conn'] = nxt
                    rev_segs.append(s)
                segs2 = rev_segs
                end_pin_final = row2['start_pin']
            
            if segs1:
                segs1[-1]['next_conn'] = ref
            
            merged_segs = segs1 + segs2
            total_len = sum(s['length'] for s in merged_segs)
            
            c1 = row1['via_count']
            c2 = row2['via_count']
            new_via_count = c1 + c2
            
            n1 = row1['net_name']
            n2 = row2['net_name']
            clean1 = get_clean_name(n1)
            clean2 = get_clean_name(n2)
            final_name = clean2 if len(clean2) < len(n2) else clean1

            new_row = {
                'source_file': row1['source_file'],
                'net_name': final_name,
                'start_pin': start_pin_final,
                'end_pin': end_pin_final,
                'segments': merged_segs,
                'via_count': new_via_count,
                'total': total_len
            }
            new_rows.append(new_row)
        
        if not merged_any:
            break
            
        final_list = [r for i, r in enumerate(rows) if i not in indices_to_remove]
        final_list.extend(new_rows)
        rows = final_list
    
    return rows


def collect_and_sort_layers(grouped_data: Dict) -> List[str]:
    """
    收集所有 Layer 並按照 TOP -> L1 -> L2... -> BOTTOM 順序排序。
    """
    all_layers = set()
    for (ref1, ref2), rows in grouped_data.items():
        for row in rows:
            for seg in row['segments']:
                layer = seg['layer']
                if layer:
                    all_layers.add(layer.strip())
    
    def layer_sort_key(name):
        u = name.upper()
        if 'TOP' in u: return -1000
        if 'BOTTOM' in u or 'BOT' in u: return 1000
        nums = re.findall(r'\d+', name)
        if nums: return int(nums[0])
        return 0
        
    sorted_layers = sorted(list(all_layers), key=layer_sort_key)
    return sorted_layers


def combine_to_excel(input_files: List[str], output_file: str, include_source: bool = True):
    if not HAS_OPENPYXL:
        raise ImportError("The 'openpyxl' library is required for Excel export. Please install it via 'pip install openpyxl'.")

    all_raw_nets = []
    for fp in input_files:
        if not os.path.exists(fp):
            print(f'[WARN] file not found: {fp}')
            continue
        nets = parse_file(fp)
        all_raw_nets.extend(nets)

    if not all_raw_nets:
        raise RuntimeError('No net resolved.')

    segment_rows = [convert_net_to_segments(n) for n in all_raw_nets]
    merged_rows = merge_passive_nets(segment_rows)
    
    # --- Grouping ---
    grouped_data = {} 
    for r in merged_rows:
        s_ref = get_refdes(r['start_pin']) or "Unknown"
        e_ref = get_refdes(r['end_pin']) or "Unknown"
        refs = sorted([s_ref, e_ref])
        key = (refs[0], refs[1])
        if key not in grouped_data:
            grouped_data[key] = []
        grouped_data[key].append(r)
    
    # --- Collect & Sort Layers ---
    sorted_layers = collect_and_sort_layers(grouped_data)
    
    # --- Write Excel ---
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    INTERFACE_OPTIONS = ['PCIe2', 'PCIe3', 'PCIe4', 'PCIe5', 'PCIe6', 'xGMI', 'UPI']
    
    START_DEVICE_DATA = {
        'CPU': {'4G': 0.5, '8G': 1.0, '16G': 2.0},
        'Switch': {'4G': 0.3, '8G': 0.6, '16G': 1.2}
    }
    
    END_DEVICE_DATA = {
        'Near stack': {'4G': 0.2, '8G': 0.4, '16G': 0.8},
        'CEM': {'4G': 0.4, '8G': 0.8, '16G': 1.6},
        'MCIO': {'4G': 0.6, '8G': 1.2, '16G': 2.4},
        'OCP conn': {'4G': 0.5, '8G': 1.0, '16G': 2.0},
        'Switch': {'4G': 0.3, '8G': 0.6, '16G': 1.2}
    }
    
    CABLE_DATA = {
        'CatA': {'4G': 10.5, '8G': 21.0, '16G': 42.0},
        'CatB': {'4G': 12.3, '8G': 24.6, '16G': 49.2},
        'CatC': {'4G': 15.0, '8G': 30.0, '16G': 60.0},
        'CatD': {'4G': 18.5, '8G': 37.0, '16G': 74.0},
        'CatE': {'4G': 22.0, '8G': 44.0, '16G': 88.0},
        'CatF': {'4G': 25.5, '8G': 51.0, '16G': 102.0}
    }
    
    LAYER_LOSS_DATA = {
        'TOP': {'4G': 2.5, '8G': 5.0, '16G': 10.0},
        'L1': {'4G': 2.0, '8G': 4.0, '16G': 8.0},
        'L2': {'4G': 1.8, '8G': 3.6, '16G': 7.2},
        'L3': {'4G': 1.8, '8G': 3.6, '16G': 7.2},
        'BOTTOM': {'4G': 2.5, '8G': 5.0, '16G': 10.0},
        'VIA': {'4G': 0.1, '8G': 0.2, '16G': 0.4},
        'Cap': {'4G': 0.05, '8G': 0.1, '16G': 0.2}
    }

    # Config Sheet setup
    config_ws = wb.create_sheet(title='Config')
    
    # Row 1 Headers
    config_ws['A1'] = "Interface_List"
    config_ws['B1'] = "Nyquist frequency (GHz)"
    config_ws['C1'] = "Risk level (Loss Limit)"
    
    # Headers Styling
    for cell in ['A1', 'B1', 'C1']:
        config_ws[cell].font = openpyxl.styles.Font(bold=True)
        config_ws[cell].alignment = openpyxl.styles.Alignment(horizontal='center')
        config_ws[cell].fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Interface options
    for i, opt in enumerate(INTERFACE_OPTIONS, start=2):
        config_ws[f'A{i}'] = opt
        # Default values for user to edit (optional, but helpful)
        config_ws[f'B{i}'] = 16  # Default 16GHz
        config_ws[f'C{i}'] = 15.0 # Default Limit
        
    interface_range = "'Config'!$A$2:$A$100"
    
    # Column Widths
    config_ws.column_dimensions['A'].width = 15
    config_ws.column_dimensions['B'].width = 25
    config_ws.column_dimensions['C'].width = 25

    # [MODIFIED] Make Config sheet visible for user input
    config_ws.sheet_state = 'visible' 
    
    # 1. Summary Sheet
    summary_ws = wb.create_sheet(title='Summary', index=0)
    
    summary_headers = [
        'Interface', 'Sheet Name', 'Component Pair', 'Net Count', 
        'Longest net', 'Longest length (mil)', 
        'Shortest net', 'Shortest length (mil)',
        'Highest loss Net', 'Loss at 4G', 'Loss at 8G', 'Loss at 16G', 'Risk Level'
    ]
    
    for idx, header_text in enumerate(summary_headers, start=1):
        cell = summary_ws.cell(row=1, column=idx, value=header_text)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))

    summary_row_idx = 2
    
    # 2. Loss data Sheet
    loss_data_ws = wb.create_sheet(title='loss data', index=1)
    
    loss_data_ws['A1'] = 'Layer'
    loss_data_ws['B1'] = '4G (loss/inch)'
    loss_data_ws['C1'] = '8G (loss/inch)'
    loss_data_ws['D1'] = '16G (loss/inch)'
    
    for col in ['A', 'B', 'C', 'D']:
        cell = loss_data_ws[f'{col}1']
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))

    loss_data_row = 2
    for layer in sorted_layers:
        loss_data_ws[f'A{loss_data_row}'] = layer
        if layer in LAYER_LOSS_DATA:
            loss_data_ws[f'B{loss_data_row}'] = LAYER_LOSS_DATA[layer]['4G']
            loss_data_ws[f'C{loss_data_row}'] = LAYER_LOSS_DATA[layer]['8G']
            loss_data_ws[f'D{loss_data_row}'] = LAYER_LOSS_DATA[layer]['16G']
        loss_data_row += 1
    
    extra_items = ['VIA', 'Cap']
    for item in extra_items:
        cell = loss_data_ws[f'A{loss_data_row}']
        cell.value = item
        if item in LAYER_LOSS_DATA:
            loss_data_ws[f'B{loss_data_row}'] = LAYER_LOSS_DATA[item]['4G']
            loss_data_ws[f'C{loss_data_row}'] = LAYER_LOSS_DATA[item]['8G']
            loss_data_ws[f'D{loss_data_row}'] = LAYER_LOSS_DATA[item]['16G']
        loss_data_row += 1
    
    loss_data_ws['F1'] = "Start Device List"
    loss_data_ws['G1'] = '4G (loss/unit)'
    loss_data_ws['H1'] = '8G (loss/unit)'
    loss_data_ws['I1'] = '16G (loss/unit)'
    
    for col in ['F', 'G', 'H', 'I']:
        cell = loss_data_ws[f'{col}1']
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
    
    start_device_row = 2
    for device, loss_values in START_DEVICE_DATA.items():
        loss_data_ws[f'F{start_device_row}'] = device
        loss_data_ws[f'G{start_device_row}'] = loss_values['4G']
        loss_data_ws[f'H{start_device_row}'] = loss_values['8G']
        loss_data_ws[f'I{start_device_row}'] = loss_values['16G']
        start_device_row += 1
    
    start_dev_range = "'loss data'!$F$2:$F$100"
    
    loss_data_ws['K1'] = "End Device List"
    loss_data_ws['L1'] = '4G (loss/unit)'
    loss_data_ws['M1'] = '8G (loss/unit)'
    loss_data_ws['N1'] = '16G (loss/unit)'
    
    for col in ['K', 'L', 'M', 'N']:
        cell = loss_data_ws[f'{col}1']
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
    
    end_device_row = 2
    for device, loss_values in END_DEVICE_DATA.items():
        loss_data_ws[f'K{end_device_row}'] = device
        loss_data_ws[f'L{end_device_row}'] = loss_values['4G']
        loss_data_ws[f'M{end_device_row}'] = loss_values['8G']
        loss_data_ws[f'N{end_device_row}'] = loss_values['16G']
        end_device_row += 1
    
    end_dev_range = "'loss data'!$K$2:$K$100"
    
    loss_data_ws['P1'] = "Cable List"
    loss_data_ws['Q1'] = '4G (loss/1000mm)'
    loss_data_ws['R1'] = '8G (loss/1000mm)'
    loss_data_ws['S1'] = '16G (loss/1000mm)'
    
    for col in ['P', 'Q', 'R', 'S']:
        cell = loss_data_ws[f'{col}1']
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
    
    cable_row = 2
    for cable, loss_values in CABLE_DATA.items():
        loss_data_ws[f'P{cable_row}'] = cable
        loss_data_ws[f'Q{cable_row}'] = loss_values['4G']
        loss_data_ws[f'R{cable_row}'] = loss_values['8G']
        loss_data_ws[f'S{cable_row}'] = loss_values['16G']
        cable_row += 1
    
    cable_range = "'loss data'!$P$2:$P$100"
    
    for col in ['A','B','C','D','F','G','H','I','K','L','M','N','P','Q','R','S']:
        loss_data_ws.column_dimensions[col].width = 18

    def should_have_loss_calc(net_name):
        name_upper = net_name.strip().upper()
        return name_upper.startswith('PE') or name_upper.startswith('XGMI') or name_upper.startswith('UPI')
        
    summary_info_list = []

    # 3. Data Sheets
    for (ref1, ref2), rows in sorted(grouped_data.items()):
        sheet_name = f"{ref1}_{ref2}"
        invalid_chars = '[]:*?/\\' 
        for ch in invalid_chars:
            sheet_name = sheet_name.replace(ch, '_')
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        
        ws = wb.create_sheet(title=sheet_name)
        ws['A1'] = "Return to Summary"
        ws['A1'].hyperlink = "#'Summary'!A1"
        ws['A1'].font = openpyxl.styles.Font(color="0563C1", underline="single")

        has_loss_calc_net = any(should_have_loss_calc(r['net_name']) for r in rows)
        
        max_segments = 0
        for r in rows:
            max_segments = max(max_segments, len(r['segments']))
        
        header = []
        if include_source: header.append('source_file')
        
        if has_loss_calc_net: header.append('Start Device')
        
        header.append('net_name')
        header.append('Start Pin')
        
        for i in range(1, max_segments + 1):
            header += [f'Layer_{i}', f'Length_{i}']
            if i < max_segments:
                header += [f'VIA_{i}']
        
        header.append('End Pin')
        
        if has_loss_calc_net: header.append('End Device')
        
        if has_loss_calc_net:
            header += ['Cable level', 'Cable length (mm)']
        
        header += ['total length', 'via count']
        
        if has_loss_calc_net:
            header += ['Loss at 4G', 'Loss at 8G', 'Loss at 16G']
        
        ws.append(header)
        
        # Capture column indices
        try:
            net_name_idx = header.index('net_name') + 1 
            total_len_idx = header.index('total length') + 1
            
            if has_loss_calc_net:
                loss4g_idx = header.index('Loss at 4G') + 1
                loss8g_idx = header.index('Loss at 8G') + 1
                loss16g_idx = header.index('Loss at 16G') + 1
                
                loss4g_col = get_column_letter(loss4g_idx)
                loss8g_col = get_column_letter(loss8g_idx)
                loss16g_col = get_column_letter(loss16g_idx)
            else:
                loss4g_col = loss8g_col = loss16g_col = None
            
            net_name_col_letter = get_column_letter(net_name_idx)
            total_len_col_letter = get_column_letter(total_len_idx)
        except ValueError:
            net_name_col_letter = 'A'
            total_len_col_letter = 'B'
            loss4g_col = loss8g_col = loss16g_col = None

        rows.sort(key=lambda x: x['net_name'])
        
        current_row_idx = 3
        
        for r in rows:
            row_data = []
            if include_source: row_data.append(r['source_file'])
            
            start_device_col_idx = -1
            if has_loss_calc_net:
                start_device_col_idx = len(row_data)
                row_data.append('CPU')
            
            row_data += [r['net_name'], r['start_pin']]
            
            segs = r['segments']
            
            base_col_idx = 4 if include_source else 3
            if has_loss_calc_net: base_col_idx += 1 
            
            formula_components_4g = []
            formula_components_8g = []
            formula_components_16g = []
            
            is_loss_net = should_have_loss_calc(r['net_name'])
            
            for i in range(max_segments):
                if i < len(segs):
                    s = segs[i]
                    row_data += [s['layer'], round(s['length'], 2)]
                    
                    if is_loss_net:
                        layer_col_letter = get_column_letter(base_col_idx + i*3)
                        len_col_letter = get_column_letter(base_col_idx + i*3 + 1)
                        
                        term_4g = f"(({len_col_letter}{current_row_idx}/1000)*VLOOKUP({layer_col_letter}{current_row_idx},'loss data'!$A:$D,2,0))"
                        term_8g = f"(({len_col_letter}{current_row_idx}/1000)*VLOOKUP({layer_col_letter}{current_row_idx},'loss data'!$A:$D,3,0))"
                        term_16g = f"(({len_col_letter}{current_row_idx}/1000)*VLOOKUP({layer_col_letter}{current_row_idx},'loss data'!$A:$D,4,0))"
                        
                        formula_components_4g.append(term_4g)
                        formula_components_8g.append(term_8g)
                        formula_components_16g.append(term_16g)
                    
                    if i < max_segments - 1:
                        conn = s['next_conn']
                        row_data += [conn if conn else '']
                else:
                    row_data += ['', '']
                    if i < max_segments - 1:
                        row_data += ['']
            
            row_data.append(r['end_pin'])
            
            end_device_col_idx = -1
            if has_loss_calc_net:
                end_device_col_idx = len(row_data)
                row_data.append('Near stack')
            
            cable_level_col_idx = -1
            cable_length_col_idx = -1
            if has_loss_calc_net:
                cable_level_col_idx = len(row_data)
                cable_length_col_idx = len(row_data) + 1
                row_data += ['CatF', 0] 
            
            row_data += [round(r['total'], 2), r['via_count']]
            
            if has_loss_calc_net:
                if is_loss_net:
                    via_cnt = r['via_count']
                    cap_cnt = 0
                    for s in segs:
                        if s['next_conn'] and s['next_conn'].startswith('C'): cap_cnt += 1
                    
                    # Start Pin Loss
                    start_device_cell = get_column_letter(start_device_col_idx + 1) if start_device_col_idx != -1 else 'B'
                    f_start_4g = f"(VLOOKUP({start_device_cell}{current_row_idx},'loss data'!$F:$I,2,0))"
                    f_start_8g = f"(VLOOKUP({start_device_cell}{current_row_idx},'loss data'!$F:$I,3,0))"
                    f_start_16g = f"(VLOOKUP({start_device_cell}{current_row_idx},'loss data'!$F:$I,4,0))"
                    
                    # End Pin Loss
                    end_device_cell = get_column_letter(end_device_col_idx + 1) if end_device_col_idx != -1 else 'C'
                    f_end_4g = f"(VLOOKUP({end_device_cell}{current_row_idx},'loss data'!$K:$N,2,0))"
                    f_end_8g = f"(VLOOKUP({end_device_cell}{current_row_idx},'loss data'!$K:$N,3,0))"
                    f_end_16g = f"(VLOOKUP({end_device_cell}{current_row_idx},'loss data'!$K:$N,4,0))"
                    
                    formula_components_4g.insert(0, f_start_4g)
                    formula_components_8g.insert(0, f_start_8g)
                    formula_components_16g.insert(0, f_start_16g)
                    
                    formula_components_4g.append(f_end_4g)
                    formula_components_8g.append(f_end_8g)
                    formula_components_16g.append(f_end_16g)
                    
                    if cable_level_col_idx != -1 and cable_length_col_idx != -1:
                        cable_level_cell = get_column_letter(cable_level_col_idx + 1)
                        cable_length_cell = get_column_letter(cable_length_col_idx + 1)
                        
                        f_cable_4g = f"(({cable_length_cell}{current_row_idx}/1000)*VLOOKUP({cable_level_cell}{current_row_idx},'loss data'!$P:$S,2,0))"
                        f_cable_8g = f"(({cable_length_cell}{current_row_idx}/1000)*VLOOKUP({cable_level_cell}{current_row_idx},'loss data'!$P:$S,3,0))"
                        f_cable_16g = f"(({cable_length_cell}{current_row_idx}/1000)*VLOOKUP({cable_level_cell}{current_row_idx},'loss data'!$P:$S,4,0))"
                        
                        formula_components_4g.append(f_cable_4g)
                        formula_components_8g.append(f_cable_8g)
                        formula_components_16g.append(f_cable_16g)
                    
                    if via_cnt > 0:
                        f_via_4g = f"({via_cnt}*VLOOKUP(\"VIA\",'loss data'!$A:$D,2,0))"
                        f_via_8g = f"({via_cnt}*VLOOKUP(\"VIA\",'loss data'!$A:$D,3,0))"
                        f_via_16g = f"({via_cnt}*VLOOKUP(\"VIA\",'loss data'!$A:$D,4,0))"
                        formula_components_4g.append(f_via_4g)
                        formula_components_8g.append(f_via_8g)
                        formula_components_16g.append(f_via_16g)
                    
                    if cap_cnt > 0:
                        f_cap_4g = f"({cap_cnt}*VLOOKUP(\"Cap\",'loss data'!$A:$D,2,0))"
                        f_cap_8g = f"({cap_cnt}*VLOOKUP(\"Cap\",'loss data'!$A:$D,3,0))"
                        f_cap_16g = f"({cap_cnt}*VLOOKUP(\"Cap\",'loss data'!$A:$D,4,0))"
                        formula_components_4g.append(f_cap_4g)
                        formula_components_8g.append(f_cap_8g)
                        formula_components_16g.append(f_cap_16g)
                    
                    final_f_4g = "=" + "+".join(formula_components_4g) if formula_components_4g else ""
                    final_f_8g = "=" + "+".join(formula_components_8g) if formula_components_8g else ""
                    final_f_16g = "=" + "+".join(formula_components_16g) if formula_components_16g else ""
                    
                    row_data += [final_f_4g, final_f_8g, final_f_16g]
                else:
                    row_data += ['', '', '']
            
            ws.append(row_data)
            
            if has_loss_calc_net:
                if start_device_col_idx != -1:
                    dv_start = DataValidation(type="list", formula1=start_dev_range, allow_blank=False)
                    dv_start.add(ws[f'{get_column_letter(start_device_col_idx + 1)}{current_row_idx}'])
                    ws.add_data_validation(dv_start)
                
                if end_device_col_idx != -1:
                    dv_end = DataValidation(type="list", formula1=end_dev_range, allow_blank=False)
                    dv_end.add(ws[f'{get_column_letter(end_device_col_idx + 1)}{current_row_idx}'])
                    ws.add_data_validation(dv_end)
                
                if cable_level_col_idx != -1:
                    dv_cable = DataValidation(type="list", formula1=cable_range, allow_blank=True)
                    dv_cable.add(ws[f'{get_column_letter(cable_level_col_idx + 1)}{current_row_idx}'])
                    ws.add_data_validation(dv_cable)
            
            current_row_idx += 1
        
        summary_info_list.append({
            'sheet_name': sheet_name,
            'ref1': ref1,
            'ref2': ref2,
            'net_count': len(rows),
            'net_col': net_name_col_letter,
            'len_col': total_len_col_letter,
            'has_loss': has_loss_calc_net,
            'loss4g_col': loss4g_col,
            'loss8g_col': loss8g_col,
            'loss16g_col': loss16g_col
        })

    # 4. Fill Summary Sheet
    for info in summary_info_list:
        sh_name = info['sheet_name']
        r = summary_row_idx # Current Summary Row
        
        summary_ws[f'A{r}'] = 'PCIe4'
        dv_interface = DataValidation(type="list", formula1=interface_range, allow_blank=False)
        dv_interface.add(summary_ws[f'A{r}'])
        summary_ws.add_data_validation(dv_interface)
        
        cell = summary_ws[f'B{r}']
        cell.value = sh_name
        cell.hyperlink = f"#'{sh_name}'!A1"
        cell.font = openpyxl.styles.Font(color="0563C1", underline="single")
        
        summary_ws[f'C{r}'] = f"{info['ref1']} <-> {info['ref2']}"
        summary_ws[f'D{r}'] = info['net_count']
        
        len_range_str = f"'{sh_name}'!{info['len_col']}:{info['len_col']}"
        name_range_str = f"'{sh_name}'!{info['net_col']}:{info['net_col']}"
        
        match_max_part = f"MATCH(MAX({len_range_str}),{len_range_str},0)"
        link_addr_max = f"\"#'{sh_name}'!A\"&{match_max_part}"
        display_name_max = f"INDEX({name_range_str},{match_max_part})"
        summary_ws[f'E{r}'].value = f'=HYPERLINK({link_addr_max}, {display_name_max})'
        summary_ws[f'E{r}'].font = openpyxl.styles.Font(color="0563C1", underline="single")

        summary_ws[f'F{r}'].value = f'=MAX({len_range_str})'
        summary_ws[f'F{r}'].number_format = '0.00' 

        match_min_part = f"MATCH(MIN({len_range_str}),{len_range_str},0)"
        link_addr_min = f"\"#'{sh_name}'!A\"&{match_min_part}"
        display_name_min = f"INDEX({name_range_str},{match_min_part})"
        summary_ws[f'G{r}'].value = f'=HYPERLINK({link_addr_min}, {display_name_min})'
        summary_ws[f'G{r}'].font = openpyxl.styles.Font(color="0563C1", underline="single")

        summary_ws[f'H{r}'].value = f'=MIN({len_range_str})'
        summary_ws[f'H{r}'].number_format = '0.00' 
        
        if info['has_loss']:
            loss16g_range = f"'{sh_name}'!{info['loss16g_col']}:{info['loss16g_col']}"
            loss8g_range = f"'{sh_name}'!{info['loss8g_col']}:{info['loss8g_col']}"
            loss4g_range = f"'{sh_name}'!{info['loss4g_col']}:{info['loss4g_col']}"
            
            match_max_loss_row = f"MATCH(MAX({loss16g_range}),{loss16g_range},0)"
            
            # Highest loss Net Name
            link_loss_addr = f"\"#'{sh_name}'!A\"&{match_max_loss_row}"
            display_loss_name = f"INDEX({name_range_str},{match_max_loss_row})"
            summary_ws[f'I{r}'].value = f'=HYPERLINK({link_loss_addr}, {display_loss_name})'
            summary_ws[f'I{r}'].font = openpyxl.styles.Font(color="0563C1", underline="single")
            
            # Loss values
            summary_ws[f'J{r}'].value = f'=INDEX({loss4g_range}, {match_max_loss_row})'
            summary_ws[f'J{r}'].number_format = '0.00'

            summary_ws[f'K{r}'].value = f'=INDEX({loss8g_range}, {match_max_loss_row})'
            summary_ws[f'K{r}'].number_format = '0.00'

            summary_ws[f'L{r}'].value = f'=MAX({loss16g_range})'
            summary_ws[f'L{r}'].number_format = '0.00'

            # [MODIFIED] Risk Level Formula based on Config
            cfg_tbl = "'Config'!$A$2:$C$100"
            # Interface is in Col A
            # Config Col B is Frequency, Col C is Limit
            
            # Logic:
            # 1. Lookup Freq Limit from Config!B based on Interface
            # 2. Lookup Loss Limit from Config!C based on Interface
            # 3. IF Freq <= 4, check Loss4G (Col J) > Limit
            # 4. IF Freq <= 8, check Loss8G (Col K) > Limit
            # 5. ELSE (Freq > 8), check Loss16G (Col L) > Limit
            
            risk_formula = (
                f'=IFERROR('
                f'IF(VLOOKUP($A{r},{cfg_tbl},2,0)<=4,'
                f'IF($J{r}>VLOOKUP($A{r},{cfg_tbl},3,0),"High","Low"),'
                f'IF(VLOOKUP($A{r},{cfg_tbl},2,0)<=8,'
                f'IF($K{r}>VLOOKUP($A{r},{cfg_tbl},3,0),"High","Low"),'
                f'IF($L{r}>VLOOKUP($A{r},{cfg_tbl},3,0),"High","Low")'
                f')'
                f')'
                f', "OK")'
            )
            
            # Use conditional formatting or color later if needed, now just text
            cell_risk = summary_ws[f'M{r}']
            cell_risk.value = risk_formula
            cell_risk.font = openpyxl.styles.Font(bold=True)
            
        summary_row_idx += 1
    
    # Adjust Summary Widths
    summary_ws.column_dimensions['A'].width = 15
    summary_ws.column_dimensions['B'].width = 30
    summary_ws.column_dimensions['C'].width = 35
    summary_ws.column_dimensions['D'].width = 12
    summary_ws.column_dimensions['E'].width = 35
    summary_ws.column_dimensions['F'].width = 20
    summary_ws.column_dimensions['G'].width = 35
    summary_ws.column_dimensions['H'].width = 20
    summary_ws.column_dimensions['I'].width = 35
    summary_ws.column_dimensions['J'].width = 15
    summary_ws.column_dimensions['K'].width = 15
    summary_ws.column_dimensions['L'].width = 15
    summary_ws.column_dimensions['M'].width = 15

    wb.save(output_file)


def run_gui():
    root = tk.Tk()
    root.withdraw()
    
    if not HAS_OPENPYXL:
        messagebox.showerror('Missing Library', "Please install 'openpyxl' to use Excel export feature.\nRun: pip install openpyxl")
        return

    messagebox.showinfo('Select File', 'Please select one or more ECL report files (txt)')
    file_paths = filedialog.askopenfilenames(
        title='Choose the ECL report file',
        filetypes=[('Text files', '*.txt'), ('All files', '*.*')]
    )
    if not file_paths: return
    
    ofile = filedialog.asksaveasfilename(
        title='output Excel file',
        defaultextension='.xlsx',
        filetypes=[('Excel files', '*.xlsx')],
        initialfile='ecl_reports_combined.xlsx'
    )
    if not ofile: return

    try:
        combine_to_excel(list(file_paths), ofile, include_source=True)
        messagebox.showinfo('Completed', f'Exported: {ofile}')
    except Exception as e:
        messagebox.showerror('Error', f'Processing failure: {e}')
        import traceback
        traceback.print_exc()


if __name__ == '__main__':
    try:
        run_gui()
    except tk.TclError:
        print('GUI not available.')
